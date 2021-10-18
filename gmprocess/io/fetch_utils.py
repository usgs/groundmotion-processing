# stdlib imports
import os
import json
import logging
import warnings
import glob
import re
import requests
from collections import OrderedDict
from datetime import datetime
from setuptools_scm import get_version

# third party imports
import folium
from libcomcat.search import get_event_by_id
from obspy.geodetics.base import locations2degrees
from obspy.taup import TauPyModel
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
import pandas as pd
from openpyxl import load_workbook
from ruamel.yaml import YAML
from ruamel.yaml.error import YAMLError
import numpy as np
from impactutils.mapping.city import Cities
from impactutils.mapping.mercatormap import MercatorMap
from impactutils.mapping.scalebar import draw_scale
from libcomcat.search import get_event_by_id
from cartopy import feature as cfeature
import prov.model

# local imports
from gmprocess.core.streamcollection import StreamCollection
from gmprocess.core.stationtrace import (
    NS_SEIS, _get_person_agent, _get_software_agent)
from gmprocess.utils.event import get_event_object, ScalarEvent
from gmprocess.utils.config import get_config, update_dict
from gmprocess.utils.constants import (
    RUPTURE_FILE, WORKSPACE_NAME, EVENT_TIMEFMT, COMPONENTS)
from gmprocess.io.asdf.stream_workspace import StreamWorkspace
from gmprocess.io.cosmos.core import BUILDING_TYPES
from gmprocess.io.global_fetcher import fetch_data
from gmprocess.io.read_directory import directory_to_streams
from gmprocess.io.stream import streams_to_dataframe
from gmprocess.metrics.station_summary import XML_UNITS

TIMEFMT2 = '%Y-%m-%dT%H:%M:%S.%f'


OCEAN_COLOR = '#96e8ff'
LAND_COLOR = '#ededaf'
PASSED_COLOR = '#00ac00'
FAILED_COLOR = '#ff2222'

MAP_PADDING = 1.1  # Station map padding value

UNITS = {
    'PGA': r'%g',
    'PGV': r'cm/s',
    'SA': r'%g'
}

FLOAT_PATTERN = r'[-+]?[0-9]*\.?[0-9]+'


def download(event, event_dir, config, directory, create_workspace=True,
             stream_collection=True):
    """Download data or load data from local directory, turn into Streams.

    Args:
        event (ScalarEvent):
            Object containing basic event hypocenter, origin time, magnitude.
        event_dir (str):
            Path where raw directory should be created (if downloading).
        config (dict):
            Dictionary with gmprocess configuration information.
        directory (str):
            Path where data already exists. Must be organized in a 'raw'
            directory, within directories with names as the event ids. For
            example, if `directory` is 'proj_dir' and you have data for
            event id 'abc123' then the raw data to be read in should be
            located in `proj_dir/abc123/raw/`.
        create_workspace (bool):
            Create workspace file?
        stream_collection (bool):
            Construct and return a StreamCollection instance?

    Returns:
        tuple:
            - StreamWorkspace: Contains the event and raw streams.
            - str: Name of workspace HDF file.
            - StreamCollection: Raw data StationStreams.
            - str: Path to the rupture file.
    """
    # Make raw directory
    rawdir = get_rawdir(event_dir)

    if directory is None:
        tcollection, terrors = fetch_data(
            event.time.datetime,
            event.latitude,
            event.longitude,
            event.depth_km,
            event.magnitude,
            config=config,
            rawdir=rawdir,
            stream_collection=stream_collection)
        # download an event.json file in each event directory,
        # in case user is simply downloading for now
        create_event_file(event, event_dir)
        download_rupture_file(event.id, event_dir)
        rup_file = get_rupture_file(event_dir)
    else:
        # Make raw directory
        in_event_dir = os.path.join(directory, event.id)
        rup_file = get_rupture_file(in_event_dir)
        in_raw_dir = get_rawdir(in_event_dir)
        logging.debug('in_raw_dir: %s' % in_raw_dir)
        streams, bad, terrors = directory_to_streams(
            in_raw_dir, config=config)
        logging.debug('streams:')
        logging.debug(streams)
        tcollection = StreamCollection(streams, **config['duplicate'])
        create_event_file(event, event_dir)
    if len(tcollection):
        logging.debug('tcollection.describe():')
        logging.debug(tcollection.describe())

    # Plot the raw waveforms
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=UserWarning)
        pngfiles = glob.glob(os.path.join(rawdir, '*.png'))
        if not len(pngfiles):
            plot_raw(rawdir, tcollection, event)

    if create_workspace:
        # Create the workspace file and put the unprocessed waveforms in it
        workname = os.path.join(event_dir, WORKSPACE_NAME)

        # Remove any existing workspace file
        if os.path.isfile(workname):
            os.remove(workname)

        workspace = StreamWorkspace(workname)
        workspace.addEvent(event)
        logging.debug('workspace.dataset.events:')
        logging.debug(workspace.dataset.events)
        workspace.addStreams(event, tcollection, label='unprocessed')
        logging.debug('workspace.dataset.waveforms.list():')
        logging.debug(workspace.dataset.waveforms.list())
    else:
        workspace = None
        workname = None
    return (workspace, workname, tcollection, rup_file)


def parse_event_file(eventfile):
    """Parse text file containing basic event information.

    Files can contain:
        - one column, in which case that column
          contains ComCat event IDs.
        - Seven columns, in which case those columns should be:
          - id: any string (no spaces)
          - time: Any ISO standard for date/time.
          - lat: Earthquake latitude in decimal degrees.
          - lon: Earthquake longitude in decimal degrees.
          - depth: Earthquake longitude in kilometers.
          - magnitude: Earthquake magnitude.
          - magnitude_type: Earthquake magnitude type.

    NB: THERE SHOULD NOT BE ANY HEADERS ON THIS FILE!

    Args:
        eventfile (str):
            Path to event text file

    Returns:
        list: ScalarEvent objects constructed from list of event information.

    """
    df = pd.read_csv(eventfile, sep=',', header=None)
    nrows, ncols = df.shape
    events = []
    if ncols == 1:
        df.columns = ['eventid']
        for idx, row in df.iterrows():
            event = get_event_object(row['eventid'])
            events.append(event)
    elif ncols == 7:
        df.columns = ['id', 'time', 'lat', 'lon', 'depth', 'magnitude',
                      'magnitude_type']
        df['time'] = pd.to_datetime(df['time'])
        for idx, row in df.iterrows():
            rowdict = row.to_dict()
            event = get_event_object(rowdict)
            events.append(event)
    else:
        return None
    return events


def draw_stations_map(pstreams, event, event_dir):

    # interactive html map is created first
    lats = np.array([stream[0].stats.coordinates['latitude']
                     for stream in pstreams])
    lons = np.array([stream[0].stats.coordinates['longitude']
                     for stream in pstreams])
    stnames = np.array([stream[0].stats.station
                        for stream in pstreams])
    networks = np.array([stream[0].stats.network
                         for stream in pstreams])

    failed = np.array([
        np.any([trace.hasParameter("failure") for trace in stream])
        for stream in pstreams])

    failure_reasons = list(pd.Series(
        [next(tr for tr in st if tr.hasParameter('failure')).
         getParameter('failure')['reason'] for st in pstreams
         if not st.passed], dtype=str))

    station_map = folium.Map(location=[event.latitude, event.longitude],
                             zoom_start=7, control_scale=True)

    failed_coords = zip(lats[failed], lons[failed])
    failed_stations = stnames[failed]
    failed_networks = networks[failed]
    failed_station_df = pd.DataFrame(
        {'stnames': failed_stations, 'network': failed_networks,
         'coords': failed_coords, 'reason': failure_reasons})

    passed_coords = zip(lats[~failed], lons[~failed])
    passed_stations = stnames[~failed]
    passed_networks = networks[~failed]
    passed_station_df = pd.DataFrame(
        {'stnames': passed_stations, 'network': passed_networks,
         'coords': passed_coords})

    # Plot the failed first
    for i, r in failed_station_df.iterrows():
        station_info = 'NET: {} LAT: {:.2f} LON: {:.2f} REASON: {}'.\
            format(r['network'], r['coords'][0], r['coords'][1], r['reason'])
        folium.CircleMarker(
            location=r['coords'],
            tooltip=r['stnames'], popup=station_info,
            color=FAILED_COLOR, fill=True, radius=6).add_to(station_map)

    for i, r in passed_station_df.iterrows():
        station_info = 'NET: {}\n LAT: {:.2f} LON: {:.2f}'.\
            format(r['network'], r['coords'][0], r['coords'][1])
        folium.CircleMarker(
            location=r['coords'], tooltip=r['stnames'], popup=station_info,
            color=PASSED_COLOR, fill=True, radius=10).add_to(station_map)

    event_info = 'MAG: {} LAT: {:.2f} LON: {:.2f} DEPTH: {:.2f}'.\
        format(event.magnitude, event.latitude, event.longitude, event.depth)
    folium.CircleMarker(
        [event.latitude, event.longitude], popup=event_info,
        color='yellow', fill=True, radius=15).add_to(station_map)

    mapfile = os.path.join(event_dir, 'stations_map.html')
    station_map.save(mapfile)

    # now the static map for the report is created
    # draw map of stations and cities and stuff
    cy = event.latitude
    cx = event.longitude
    xmin = lons.min()
    xmax = lons.max()
    ymin = lats.min()
    ymax = lats.max()

    diff_x = max(abs(cx - xmin), abs(cx - xmax), 1)
    diff_y = max(abs(cy - ymin), abs(cy - ymax), 1)

    xmax = cx + MAP_PADDING * diff_x
    xmin = cx - MAP_PADDING * diff_x
    ymax = cy + MAP_PADDING * diff_y
    ymin = cy - MAP_PADDING * diff_y

    bounds = (xmin, xmax, ymin, ymax)
    figsize = (10, 10)
    cities = Cities.fromDefault()
    mmap = MercatorMap(bounds, figsize, cities)
    mmap.drawCities(draw_dots=True)
    ax = mmap.axes
    draw_scale(ax)
    ax.plot(cx, cy, 'r*', markersize=16,
            transform=mmap.geoproj, zorder=8)

    failed = np.array([
        np.any([trace.hasParameter("failure") for trace in stream])
        for stream in pstreams])

    # Plot the failed first
    ax.scatter(lons[failed], lats[failed], c=FAILED_COLOR,
               marker='v', edgecolors='k', transform=mmap.geoproj, zorder=100,
               s=48)

    # Plot the successes above the failures
    ax.scatter(lons[~failed], lats[~failed], c=PASSED_COLOR,
               marker='^', edgecolors='k', transform=mmap.geoproj, zorder=101,
               s=48)

    passed_marker = mlines.Line2D(
        [], [], color=PASSED_COLOR, marker='^',
        markeredgecolor='k', markersize=12,
        label='Passed station', linestyle='None')
    failed_marker = mlines.Line2D(
        [], [], color=FAILED_COLOR, marker='v',
        markeredgecolor='k', markersize=12,
        label='Failed station', linestyle='None')
    earthquake_marker = mlines.Line2D(
        [], [], color='red', marker='*',
        markersize=12,
        label='Earthquake Epicenter',
        linestyle='None')
    ax.legend(handles=[passed_marker, failed_marker, earthquake_marker],
              fontsize=12)

    scale = '50m'
    land = cfeature.NaturalEarthFeature(
        category='physical',
        name='land',
        scale=scale,
        facecolor=LAND_COLOR)
    ocean = cfeature.NaturalEarthFeature(
        category='physical',
        name='ocean',
        scale=scale,
        facecolor=OCEAN_COLOR)
    ax.add_feature(land)
    ax.add_feature(ocean)
    ax.coastlines(resolution=scale, zorder=10, linewidth=1)
    mapfile = os.path.join(event_dir, 'stations_map.png')
    plt.savefig(mapfile)
    return mapfile


def get_event_files(directory):
    """Get list of event.json files found underneath a data directory.

    Args:
        directory (str):
            Path to directory containing input raw data, where
            subdirectories must be event directories containing
            event.json files, where the id in that file matches
            the directory under which it is found.
    Returns:
        List of event.json files.
    """
    eventfiles = []
    for root, dirs, files in os.walk(directory):
        for name in files:
            if name == 'event.json':
                fullname = os.path.join(root, name)
                eventfiles.append(fullname)
    return eventfiles


def read_event_json_files(eventfiles):
    """Read event.json file and return ScalarEvent object.

    Args:
        eventfiles (list):
            Event.json files to be read.
    Returns:
        list: ScalarEvent objects.

    """
    events = []
    for eventfile in eventfiles:
        with open(eventfile, 'rt', encoding='utf-8') as f:

            event = json.load(f)

            try:
                origintime = datetime.fromtimestamp(
                    event["properties"]["time"] / 1000.0)
                evdict = {
                    "id": event["id"],
                    "time": origintime.strftime("%Y-%m-%dT%H:%M:%S.%f"),
                    "lat": event["geometry"]["coordinates"][1],
                    "lon": event["geometry"]["coordinates"][0],
                    "depth": event["geometry"]["coordinates"][2],
                    "magnitude": event["properties"]["mag"],
                    "magnitude_type": event["properties"]["magType"],
                }
                event = get_event_object(evdict)

            except BaseException:
                event = get_event_object(event)

            events.append(event)
    return events


def get_events(eventids, textfile, eventinfo, directory,
               outdir=None):
    """Find the list of events.

    Args:
        eventids (list or None):
            List of ComCat event IDs.
        textfile (str or None):
            Path to text file containing event IDs or info.
        eventinfo (list or None):
            List containing:
                - id Any string, no spaces.
                - time Any ISO-compatible date/time string.
                - latitude Latitude in decimal degrees.
                - longitude Longitude in decimal degrees.
                - depth Depth in kilometers.
                - magnitude Earthquake magnitude.
                - magnitude_type Earthquake magnitude type.
        directory (str):
            Path to a directory containing event subdirectories, each
            containing an event.json file, where the ID in the json file
            matches the subdirectory containing it.
        outdir (str):
            Output directory.

    Returns:
        list: ScalarEvent objects.

    """
    events = []
    if eventids is not None:
        # Get list of events from directory if it has been provided
        tevents = []
        if directory is not None:
            tevents = events_from_directory(directory)
        elif outdir is not None:
            tevents = events_from_directory(outdir)
        for eventid in eventids:
            if len(tevents) and eventid in tevents:
                event = [e for e in tevents if e.id == eventid][0]
            else:
                # This connects to comcat to get event, does not check for a
                # local json file
                event = get_event_object(eventid)
            events.append(event)
    elif textfile is not None:
        events = parse_event_file(textfile)
    elif eventinfo is not None:
        eid = eventinfo[0]
        time = eventinfo[1]
        lat = float(eventinfo[2])
        lon = float(eventinfo[3])
        dep = float(eventinfo[4])
        mag = float(eventinfo[5])
        mag_type = str(eventinfo[6])
        event = ScalarEvent()
        event.fromParams(eid, time, lat, lon, dep, mag, mag_type)
        events = [event]
    elif directory is not None:
        events = events_from_directory(directory)
    elif outdir is not None:
        events = events_from_directory(outdir)
    return events


def events_from_directory(dir):
    events = []
    eventfiles = get_event_files(dir)
    if len(eventfiles):
        events = read_event_json_files(eventfiles)
    else:
        eventids = [f for f in os.listdir(dir) if not f.startswith('.')]
        for eventid in eventids:
            try:
                event = get_event_object(eventid)
                events.append(event)

                # If the event ID has been updated, make sure to rename
                # the source folder and issue a warning to the user
                if event.id != eventid:
                    old_dir = os.path.join(dir, eventid)
                    new_dir = os.path.join(dir, event.id)
                    os.rename(old_dir, new_dir)
                    logging.warn('Directory %s has been renamed to %s.' %
                                 (old_dir, new_dir))
            except BaseException:
                logging.warning(
                    'Could not get info for event id: %s' % eventid)

    return events


def create_event_file(event, event_dir):
    """Write event.json file in event_dir.

    Args:
        event (ScalarEvent):
            Input event object.
        event_dir (str):
            Directory where event.json should be written.
    """

    # download event.json for event
    eventid = event.origins[-1].resource_id.id
    event = get_event_by_id(eventid)
    req = requests.get(event.detail_url)
    data = json.loads(req.text)

    # dump the event.json file to the event directory
    eventfile = os.path.join(event_dir, 'event.json')
    with open(eventfile, 'w') as f:
        json.dump(data, f)


def get_rawdir(event_dir):
    """Find or create raw directory if necessary.

    Args:
        event_dir (str):
            Directory where raw directory will be found or created.
    """
    rawdir = os.path.join(event_dir, 'raw')
    if not os.path.exists(rawdir):
        os.makedirs(rawdir)
    return rawdir


def save_shakemap_amps(processed, event, event_dir):
    """Write ShakeMap peak amplitudes to Excel file and ShakeMap JSON.

    Args:
        processed (StreamCollection):
            Processed waveforms.
        event (ScalarEvent):
            Event object.
        event_dir (str):
            Directory where peak amps should be written.

    Returns:
        str: Path to output amps spreadsheet.
    """
    ampfile_name = None
    if processed.n_passed:
        dataframe = streams_to_dataframe(
            processed, event=event)
        ampfile_name = os.path.join(event_dir, 'shakemap.xlsx')

        # saving with index=False not supported by pandas
        dataframe.to_excel(ampfile_name)

        wb = load_workbook(ampfile_name)
        ws = wb.active

        # we don't need the index column, so we'll delete it here
        ws.delete_cols(1)
        ws.insert_rows(1)
        ws['A1'] = 'REFERENCE'
        ws['B1'] = dataframe['SOURCE'].iloc[0]
        # somehow pandas inserted an extra row between sub-headings and
        # the beginning of the data. Delete that row.
        ws.delete_rows(4)
        wb.save(ampfile_name)

        # get shakemap json, save to output directory
        jsonfile = os.path.join(event_dir, 'gmprocess_dat.json')
        jsonstr = get_shakemap_json(dataframe)
        with open(jsonfile, 'wt', encoding='utf-8') as fp:
            fp.write(jsonstr)

    return (ampfile_name, jsonfile)


def create_json(workspace, event, event_dir, label, config=None,
                expanded_imts=False):
    """Create JSON file for ground motion parametric data.

    Args:
        workspace (StreamWorkspace):
            gmrpocess StreamWorkspace object.
        event (ScalarEvent):
            Event object.
        event_dir (str):
            Event directory.
        label (str):
            Processing label.
        config (dict):
            Configuration options.
        expanded_imts (bool):
            Use expanded IMTs. Currently this only means all the SA that have
            been computed, plus PGA and PGV (if computed). Could eventually
            expand for other IMTs also.
    """
    features = []

    station_features = []

    streams = workspace.getStreams(event.id, labels=[label], config=config)
    npassed = 0
    for stream in streams:
        if stream.passed:
            npassed += 1
    if not npassed:
        logging.info('No strong motion data found that passes tests. Exiting.')
        return (None, None, 0)

    # Creating a new provenance document and filling in the software
    # information for every trace can be slow, so here we create a
    # base provenance document that will be copied and used as a template
    base_prov = prov.model.ProvDocument()
    base_prov.add_namespace(*NS_SEIS)
    base_prov = _get_person_agent(base_prov, config)
    base_prov = _get_software_agent(base_prov)

    nfeatures = 0
    for stream in streams:
        if not stream.passed:
            continue

        # Station is the feature, and properties contain
        # channel dictionary with all information about the metrics
        feature = OrderedDict()
        properties = OrderedDict()
        properties['network_code'] = stream[0].stats.network
        properties['station_code'] = stream[0].stats.station
        # properties['location_code'] = stream[0].stats.location
        properties['name'] = stream[0].stats.standard['station_name']
        properties['provider'] = stream[0].stats.standard['source']
        properties['instrument'] = stream[0].stats.standard['instrument']
        properties['source_format'] = stream[0].stats.standard['source_format']
        struct_desc = stream[0].stats.standard['structure_type']
        struct_type = _get_cosmos_code(struct_desc)
        properties['station_housing'] = {
            'cosmos_code': struct_type,
            'description': struct_desc
        }
        nfeatures += 1

        metrics = workspace.getStreamMetrics(
            event.id,
            properties['network_code'],
            properties['station_code'],
            label,
            config=config
        )

        if metrics is None:
            continue

        coordinates = [stream[0].stats.coordinates.longitude,
                       stream[0].stats.coordinates.latitude,
                       stream[0].stats.coordinates.elevation]

        station_feature = get_station_feature(
            stream, metrics, coordinates, expanded_imts=expanded_imts)
        if station_feature is not None:
            station_features.append(station_feature)

        components = get_components(metrics, stream)
        properties['components'] = components

        provenance = {}

        for trace in stream:
            channel = trace.stats.channel

            # get trace provenance
            provthing = trace.getProvenanceDocument(base_prov=base_prov)
            provjson = provthing.serialize(format='json')
            provenance_dict = json.loads(provjson)
            provenance[channel] = provenance_dict

        properties['provenance'] = provenance
        feature['geometry'] = {
            'type': 'Point',
            'coordinates': coordinates
        }
        feature['type'] = 'Feature'

        properties = replace_nan(properties)

        feature['properties'] = properties
        features.append(feature)

    event_dict = {
        'id': event.id,
        'time': event.time.strftime(EVENT_TIMEFMT),
        'location': '',
        'latitude': event.latitude,
        'longitude': event.longitude,
        'depth': event.depth,
        'magnitude': event.magnitude,
    }
    feature_dict = {
        'type': 'FeatureCollection',
        'software': {
            'name': 'gmprocess',
            'version': get_version(
                root=os.path.join(os.pardir, os.pardir),
                relative_to=__file__)
        },
        'process_time': datetime.utcnow().strftime(EVENT_TIMEFMT) + 'Z',
        'event': event_dict,
        'features': features
    }

    station_feature_dict = {
        'type': 'FeatureCollection',
        'features': station_features
    }
    stationfile = os.path.join(
        event_dir, '%s_groundmotions_dat.json' % event.id)
    with open(stationfile, 'wt') as f:
        json.dump(station_feature_dict, f, allow_nan=False)

    jsonfile = os.path.join(event_dir, '%s_metrics.json' % event.id)
    with open(jsonfile, 'wt') as f:
        json.dump(feature_dict, f, allow_nan=False)

    return (jsonfile, stationfile, nfeatures)


def _get_cosmos_code(desc):
    rev_types = dict(map(reversed, BUILDING_TYPES.items()))
    if desc in rev_types:
        return rev_types[desc]
    else:
        return 51


def get_shakemap_json(dataframe):
    json_dict = {'type': 'FeatureCollection'}
    features = []
    for idx, row in dataframe.iterrows():
        feature = {'type': 'Feature'}
        # the columns without a sub-header need .iloc to reference value
        lon = row['LON'].iloc[0]
        lat = row['LAT'].iloc[0]
        station = row['STATION'].iloc[0]
        network = row["NETID"].iloc[0]
        name = row["NAME"].iloc[0]
        source = row["SOURCE"].iloc[0]
        geometry = {
            'type': 'Point',
            'coordinates': (lon, lat)
        }
        sid = f'{network}.{station}'
        feature['id'] = sid
        feature['geometry'] = geometry
        props = {}
        props['code'] = station
        props['name'] = name
        props['source'] = source
        props['network'] = network

        channeldict = {}
        channels = []
        channeldict['name'] = 'H1'
        amplitudes = []
        for ampcol, value in row['GREATER_OF_TWO_HORIZONTALS'].iteritems():
            ampdict = {}
            imtname = ampcol.lower()
            if 'SA' in ampcol:
                # pull apart SA name, set period to 1 digit precision
                period = float(re.search(FLOAT_PATTERN, imtname).group())
                imtname = f'sa({period:.1f})'
                units = UNITS['SA']
            else:
                units = UNITS[ampcol]
            ampdict['name'] = imtname
            ampdict['value'] = value
            ampdict['units'] = units
            ampdict['ln_sigma'] = 0.0
            ampdict['flag'] = 0
            amplitudes.append(ampdict)
        channeldict['amplitudes'] = amplitudes
        channels.append(channeldict)

        props['channels'] = channels
        feature['properties'] = props
        features.append(feature)
    json_dict['features'] = features
    json_str = json.dumps(json_dict)
    return json_str


def update_config(custom_cfg_file):
    """Merge custom config with default.

    Args:
        custom_cfg_file (str):
            Path to custom config.

    Returns:
        dict: Merged config dictionary.

    """
    config = get_config()

    if not os.path.isfile(custom_cfg_file):
        return config
    try:
        with open(custom_cfg_file, 'rt', encoding='utf-8') as f:
            yaml = YAML()
            yaml.preserve_quotes = True
            custom_cfg = yaml.load(f)
            update_dict(config, custom_cfg)
    except YAMLError:
        return None

    return config


def plot_raw(rawdir, tcollection, event):
    """Make PNG plots of a collection of raw waveforms.

    Args:
        rawdir (str):
            Directory where PNG files should be saved.
        tcollection (StreamCollection):
            Sequence of streams.
        event (ScalarEvent):
            Event object.

    """
    model = TauPyModel(model="iasp91")
    source_depth = event.depth_km
    if source_depth < 0:
        source_depth = 0
    eqlat = event.latitude
    eqlon = event.longitude
    for stream in tcollection:
        stlat = stream[0].stats.coordinates['latitude']
        stlon = stream[0].stats.coordinates['longitude']
        dist = float(locations2degrees(eqlat, eqlon, stlat, stlon))
        try:
            arrivals = model.get_travel_times(
                source_depth_in_km=source_depth,
                distance_in_degree=dist,
                phase_list=['P', 'p', 'Pn'])
            arrival = arrivals[0]
            arrival_time = arrival.time
        except BaseException as e:
            fmt = ('Exception "%s" generated by get_travel_times() dist=%.3f '
                   'depth=%.1f')
            logging.warning(fmt % (str(e), dist, source_depth))
            arrival_time = 0.0
        ptime = arrival_time + (event.time - stream[0].stats.starttime)
        outfile = os.path.join(rawdir, '%s.png' % stream.get_id())

        fig, axeslist = plt.subplots(nrows=3, ncols=1, figsize=(12, 6))
        for ax, trace in zip(axeslist, stream):
            times = np.linspace(
                0.0, trace.stats.endtime - trace.stats.starttime,
                trace.stats.npts)
            ax.plot(times, trace.data, color='k')
            ax.set_xlabel('seconds since start of trace')
            ax.set_title('')
            ax.axvline(ptime, color='r')
            ax.set_xlim(left=0, right=times[-1])
            legstr = '%s.%s.%s.%s' % (
                trace.stats.network,
                trace.stats.station,
                trace.stats.location,
                trace.stats.channel)
            ax.legend(labels=[legstr], frameon=True, loc='upper left')
            tbefore = event.time + arrival_time < trace.stats.starttime + 1.0
            tafter = event.time + arrival_time > trace.stats.endtime - 1.0
            if tbefore or tafter:
                legstr = 'P arrival time %.1f seconds' % ptime
                left, right = ax.get_xlim()
                xloc = left + (right - left) / 20
                bottom, top = ax.get_ylim()
                yloc = bottom + (top - bottom) / 10
                ax.text(xloc, yloc, legstr, color='r')
        plt.savefig(outfile, bbox_inches='tight')
        plt.close()


def download_rupture_file(event_id, event_dir):
    """Downlaod rupture file from Comcat.

    Args:
        event_id (str):
            Event id.
        event_dir (str):
            Event directory.
    """
    event = get_event_by_id(event_id)
    try:
        shakemap_prod = event.getProducts('shakemap')
        shakemap_prod[0].getContent(
            'rupture.json', os.path.join(event_dir, 'rupture.json'))
    except BaseException:
        logging.info('%s does not have a rupture.json file.' % event_id)


def get_rupture_file(event_dir):
    """Get the path to the rupture file, or None if there is not rupture file.

    Args:
        event_dir (str):
            Event directory.

    Returns:
        str: Path to the rupture file. Returns None if no rupture file exists.
    """
    rupture_file = os.path.join(event_dir, RUPTURE_FILE)
    if not os.path.exists(rupture_file):
        rupture_file = None
    return rupture_file


def get_station_feature(stream, metrics, coordinates,
                        expanded_imts=False):
    scode = f'{stream[0].stats.network}.{stream[0].stats.station}'
    station_feature = OrderedDict()
    station_properties = OrderedDict()
    station_feature['type'] = 'Feature'
    station_feature['id'] = scode
    station_properties['name'] = stream[0].stats.standard['station_name']

    station_properties['code'] = stream[0].stats.station
    station_properties['network'] = stream[0].stats.network
    station_properties['distance'] = metrics.distances['epicentral']
    # station_properties['source'] = stream[0].stats.standard['source']
    station_properties['source'] = stream[0].stats.network
    station_channels = []
    station_channel_names = ['H1', 'H2', 'Z']

    if expanded_imts:
        imts = list(
            set([i[0] for i in metrics.pgms.index.to_numpy()
                 if i[0].startswith('SA')])
        )
        imt_lower = [s.lower() for s in imts]
        imt_units = [UNITS['SA']] * len(imts)
        if 'PGA' in metrics.pgms.index:
            imts.append('PGA')
            imt_lower.append('pga')
            imt_units.append(UNITS['PGA'])
        if 'PGV' in metrics.pgms.index:
            imts.append('PGV')
            imt_lower.append('pgv')
            imt_units.append(UNITS['PGV'])
        station_amps = {k: v for k, v in zip(imts, zip(imt_lower, imt_units))}
    else:
        station_amps = {
            'SA(0.300)': ('sa(0.3)', UNITS['SA']),
            'SA(1.000)': ('sa(1.0)', UNITS['SA']),
            'SA(3.000)': ('sa(3.0)', UNITS['SA']),
            'PGA': ('pga', UNITS['PGA']),
            'PGV': ('pgv', UNITS['PGV'])
        }

    channel_dict = metrics.channel_dict

    for channel_name in station_channel_names:
        station_channel = OrderedDict()
        if channel_name in metrics.components:
            station_channel['name'] = channel_dict[channel_name]
            station_amplitudes = []
            for gm_imt, station_tuple in station_amps.items():
                imt_value = metrics.get_pgm(gm_imt, channel_name)
                station_amplitude = OrderedDict()
                station_amplitude['name'] = station_tuple[0]
                station_amplitude['ln_sigma'] = 0
                station_amplitude['flag'] = 0
                station_amplitude['value'] = imt_value
                station_amplitude['units'] = station_tuple[1]
                station_amplitudes.append(station_amplitude.copy())
            station_channel['amplitudes'] = station_amplitudes
            station_channels.append(station_channel)
    if len(station_channels):
        station_properties['channels'] = station_channels
    else:
        return None
    station_feature['properties'] = station_properties
    station_feature['geometry'] = {
        'type': 'Point',
        'coordinates': coordinates
    }
    return station_feature


def get_components(metrics, stream):
    FLOAT_MATCH = r'[0-9]*\.[0-9]*'
    components = OrderedDict()
    for imc in metrics.components:
        if imc in ['H1', 'H2', 'Z']:
            imtlist = COMPONENTS['CHANNELS']
        else:
            imtlist = COMPONENTS[imc]
        measures = OrderedDict()
        spectral_values = []
        spectral_periods = []
        fourier_amplitudes = []
        fourier_periods = []
        for imt in metrics.imts:
            if imt.startswith('FAS'):
                imtstr = 'FAS'
            elif imt.startswith('SA'):
                imtstr = 'SA'
            else:
                imtstr = imt
            if imtstr not in imtlist:
                continue
            imt_value = metrics.get_pgm(imt, imc)
            if np.isnan(imt_value):
                imt_value = 'null'
            if imt.startswith('SA'):
                period = float(re.search(FLOAT_MATCH, imt).group())
                spectral_values.append(imt_value)
                spectral_periods.append(period)
            elif imt.startswith('FAS'):
                period = float(re.search(FLOAT_MATCH, imt).group())
                fourier_amplitudes.append(imt_value)
                fourier_periods.append(period)
            elif imt.startswith('DURATION'):
                # TODO - Make interval something retrievable from metrics
                units = XML_UNITS[imt.lower()]
                measures[imt] = {
                    'value': imt_value,
                    'units': units,
                    'interval': '5-95'
                }
            else:
                units = XML_UNITS[imt.lower()]
                measures[imt] = {'value': imt_value, 'units': units}

        if imc in ['H1', 'H2', 'Z']:
            imcname = metrics.channel_dict[imc]
            measures['as_recorded'] = True
            ttrace = stream.select(component=imcname[-1])
            azimuth = np.nan
            dip = np.nan
            if len(ttrace):
                trace = ttrace[0]
                sampling_rate = trace.stats.sampling_rate
                location_code = trace.stats.location
                peak_acc = trace.data.max()
                start = trace.stats.starttime
                delta = trace.stats.delta
                idx = np.where([trace.data >= peak_acc])[1][0]
                peak_pga_time = (start + (delta * idx)).strftime(EVENT_TIMEFMT)
                vel_trace = trace.copy()
                vel_trace.integrate()
                peak_vel = vel_trace.data.max()
                start = vel_trace.stats.starttime
                delta = vel_trace.stats.delta
                idx = np.where([vel_trace.data >= peak_vel])[1][0]
                peak_pgv_time = (start + (delta * idx)).strftime(EVENT_TIMEFMT)
                if 'horizontal_orientation' in trace.stats.standard:
                    azimuth = trace.stats.standard['horizontal_orientation']
                dip = trace.stats.standard['vertical_orientation']
            else:
                sampling_rate = np.nan
                location_code = ''
                peak_pga_time = np.nan
                peak_pgv_time = np.nan

            measures['samples_per_second'] = sampling_rate
            measures['location_code'] = location_code
            measures['peak_pga_time'] = peak_pga_time
            measures['peak_pgv_time'] = peak_pgv_time
            measures['azimuth'] = azimuth
            measures['dip'] = dip
        else:
            imcname = imc
            measures['as_recorded'] = False
        components[imcname] = measures
        if len(spectral_values):
            units = XML_UNITS['sa']
            damping = metrics.damping
            sa_dict = {
                'units': units,
                'damping': damping,
                'method': 'absolute'
            }
            sa_dict['values'] = spectral_values
            sa_dict['periods'] = spectral_periods
            components[imcname]['SA'] = sa_dict
        if len(fourier_amplitudes):
            units = XML_UNITS['fas']
            fas_dict = {
                'units': units,
                'values': fourier_amplitudes,
                'periods': fourier_periods
            }
            components[imcname]['FAS'] = fas_dict
    return components


def replace_nan(properties):
    # replace nans in any field in input dictionary with a "null" string.
    for key, value in properties.items():
        if isinstance(value, (float, np.floating)):
            if np.isnan(value):
                properties[key] = 'null'
        elif isinstance(value, dict):
            properties[key] = replace_nan(value)
    return properties
